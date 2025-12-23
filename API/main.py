from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from datetime import datetime, timedelta, timezone
from typing import Optional
from contextlib import asynccontextmanager
from database import get_cursor, init_pool, close_pool, check_database_exists
from config import DB
from pydantic import BaseModel
import hashlib
import csv
import io
import time

# Helper untuk init database
def init_db():
    """Inisialisasi database dan tabel users jika belum ada"""
    try:
        with get_cursor() as cur:
            # Create users table if not exists
            cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id SERIAL PRIMARY KEY,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    role TEXT NOT NULL DEFAULT 'user',
                    is_active BOOLEAN DEFAULT true,
                    created_at TIMESTAMP DEFAULT NOW()
                );
            """)
            
            # Check for missing columns (migration for existing tables)
            cur.execute("SELECT column_name FROM information_schema.columns WHERE table_name='users'")
            columns = [row['column_name'] for row in cur.fetchall()]
            
            if 'is_active' not in columns:
                cur.execute("ALTER TABLE users ADD COLUMN is_active BOOLEAN DEFAULT true")
            if 'created_at' not in columns:
                cur.execute("ALTER TABLE users ADD COLUMN created_at TIMESTAMP DEFAULT NOW()")
            if 'force_password_change' not in columns:
                cur.execute("ALTER TABLE users ADD COLUMN force_password_change BOOLEAN DEFAULT false")
                
            print("Database initialized successfully.")

            # Create status_relay table
            cur.execute("""
                CREATE TABLE IF NOT EXISTS status_relay (
                    id SERIAL PRIMARY KEY,
                    name TEXT NOT NULL,
                    gpio INT,
                    is_active BOOLEAN DEFAULT false
                );
            """)
            
            # Create settings table for threshold configurations
            cur.execute("""
                CREATE TABLE IF NOT EXISTS app_settings (
                    id SERIAL PRIMARY KEY,
                    setting_key TEXT UNIQUE NOT NULL,
                    setting_value JSONB NOT NULL,
                    updated_at TIMESTAMP DEFAULT NOW()
                );
            """)
            
            # Check defaults for status_relay
            cur.execute("SELECT COUNT(*) as count FROM status_relay")
            if cur.fetchone()['count'] == 0:
                 print("Seeding default status_relay...")
                 cur.execute("""
                    INSERT INTO status_relay (id, name, gpio, is_active) VALUES 
                    (1, 'Lampu Teras', 12, false),
                    (2, 'Pompa Air', 14, false),
                    (3, 'Exhaust Fan', 27, false),
                    (4, 'Door Lock', 26, false)
                 """)
            
            # Check defaults for app_settings (thresholds)
            cur.execute("SELECT COUNT(*) as count FROM app_settings WHERE setting_key = 'thresholds'")
            if cur.fetchone()['count'] == 0:
                print("Seeding default threshold settings...")
                default_thresholds = {
                    "dht22": {"tempMax": 35, "tempMin": 15, "humMax": 80, "humMin": 30},
                    "mq2": {"smokeMax": 500, "smokeWarn": 350, "lpgMax": 1000, "lpgWarn": 500, "coMax": 500, "coWarn": 200},
                    "pzem004t": {"powerMax": 2000, "voltageMin": 180, "voltageMax": 240, "currentMax": 10, "energyMax": 100, "pfMin": 0.85},
                    "bh1750": {"luxMax": 100000, "luxMin": 0}
                }
                import json
                cur.execute(
                    "INSERT INTO app_settings (setting_key, setting_value) VALUES (%s, %s)",
                    ('thresholds', json.dumps(default_thresholds))
                )
    except Exception as e:
        print(f"Database init error: {e}")

# Lifecycle events untuk connection pool
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Cek database sebelum startup
    if not check_database_exists():
        print(f"\n[FATAL] Database '{DB['dbname']}' belum ada.")
        print(f"Silakan buat database '{DB['dbname']}' terlebih dahulu.\n")
        raise RuntimeError(f"Database '{DB['dbname']}' belum dibuat.")

    # Startup: inisialisasi pool dan db
    init_pool(minconn=2, maxconn=10)
    
    # Tunggu sebentar agar DB siap (opsional, untuk docker-compose)
    # time.sleep(2) 
    
    init_db()
    yield
    # Shutdown: tutup pool
    close_pool()

app = FastAPI(title="IoT Sensor API", version="1.0", lifespan=lifespan)

origins = ["*"]  # Allow all origins for development

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# =========================================
# AUTH MODELS & HELPERS
# =========================================
class UserLogin(BaseModel):
    username: str
    password: str

class UserRegister(BaseModel):
    username: str
    password: str
    role: str = "admin"

class UserUpdate(BaseModel):
    user_id: int
    current_password: str
    username: Optional[str] = None
    new_password: Optional[str] = None

class UserCreateAdmin(BaseModel):
    username: str
    password: str
    role: str = "user"

def hash_password(password: str) -> str:
    """Hash password using SHA256"""
    return hashlib.sha256(password.encode()).hexdigest()

# =========================================
# AUTH ENDPOINTS
# =========================================
@app.get("/relays")
def get_relays():
    """Get all relays status"""
    try:
        with get_cursor() as cur:
            cur.execute("SELECT * FROM status_relay ORDER BY id ASC")
            return cur.fetchall()
    except Exception as e:
        raise HTTPException(500, f"Error: {e}")


class RelayUpdate(BaseModel):
    is_active: bool


class RelayRename(BaseModel):
    name: str


@app.put("/relays/{relay_id}")
def update_relay_status(relay_id: int, update: RelayUpdate):
    """Update single relay status (on/off)"""
    try:
        with get_cursor() as cur:
            cur.execute(
                "UPDATE status_relay SET is_active = %s WHERE id = %s RETURNING id, name, is_active",
                (update.is_active, relay_id)
            )
            result = cur.fetchone()
            if not result:
                raise HTTPException(404, f"Relay dengan ID {relay_id} tidak ditemukan")
            return {"success": True, "relay": result}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error updating relay: {e}")


@app.patch("/relays/{relay_id}/name")
def rename_relay(relay_id: int, update: RelayRename):
    """Rename relay"""
    try:
        with get_cursor() as cur:
            cur.execute(
                "UPDATE status_relay SET name = %s WHERE id = %s RETURNING id, name, gpio, is_active",
                (update.name, relay_id)
            )
            result = cur.fetchone()
            if not result:
                raise HTTPException(404, f"Relay dengan ID {relay_id} tidak ditemukan")
            return {"success": True, "relay": result}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error renaming relay: {e}")


@app.get("/relays/{relay_id}")
def get_relay_by_id(relay_id: int):
    """Get single relay by ID"""
    try:
        with get_cursor() as cur:
            cur.execute("SELECT * FROM status_relay WHERE id = %s", (relay_id,))
            result = cur.fetchone()
            if not result:
                raise HTTPException(404, f"Relay dengan ID {relay_id} tidak ditemukan")
            return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error: {e}")

@app.get("/auth/check-admin")
def check_admin():
    """Check if any admin user exists"""
    try:
        with get_cursor() as cur:
            cur.execute("SELECT COUNT(*) as count FROM users WHERE role = 'admin'")
            result = cur.fetchone()
            return {"hasAdmin": result["count"] > 0}
    except Exception as e:
        # Table might not exist yet
        return {"hasAdmin": False}

@app.post("/auth/register")
def register_user(user: UserRegister):
    """Register new user (Public: First user = Admin, others = User)"""
    try:
        with get_cursor() as cur:
            # Check if any user exists (Bootstrap Admin)
            cur.execute("SELECT COUNT(*) as count FROM users")
            count = cur.fetchone()['count']
            
            determined_role = 'admin' if count == 0 else 'user'
            
            # Check if username exists
            cur.execute("SELECT id FROM users WHERE username = %s", (user.username,))
            if cur.fetchone():
                raise HTTPException(400, "Username sudah digunakan")
            
            # Hash password and insert
            password_hash = hash_password(user.password)
            cur.execute("""
                INSERT INTO users (username, password_hash, role)
                VALUES (%s, %s, %s)
                RETURNING id, username, role
            """, (user.username, password_hash, determined_role))
            
            new_user = cur.fetchone()
            return {
                "success": True,
                "user": {
                    "id": new_user["id"],
                    "username": new_user["username"],
                    "role": new_user["role"]
                }
            }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Registration error: {e}")

# =========================================
# ADMIN USER MANAGEMENT
# =========================================
@app.get("/admin/users")
def get_all_users():
    """List all users (Admin only)"""
    try:
        with get_cursor() as cur:
            cur.execute("SELECT id, username, role, is_active, created_at FROM users ORDER BY id ASC")
            users = cur.fetchall()
            return {"users": users}
    except Exception as e:
        raise HTTPException(500, f"Error fetching users: {e}")

@app.post("/admin/users")
def create_user_by_admin(user: UserCreateAdmin):
    """Create user/admin manually (Admin only)"""
    try:
        with get_cursor() as cur:
            # Check username
            cur.execute("SELECT id FROM users WHERE username = %s", (user.username,))
            if cur.fetchone():
                raise HTTPException(400, "Username sudah digunakan")
            
            password_hash = hash_password(user.password)
            cur.execute("""
                INSERT INTO users (username, password_hash, role)
                VALUES (%s, %s, %s)
                RETURNING id, username, role, is_active
            """, (user.username, password_hash, user.role))
            
            new_user = cur.fetchone()
            return {"success": True, "user": new_user}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error creating user: {e}")

@app.delete("/admin/users/{user_id}")
def delete_user(user_id: int):
    """Delete user (Admin only)"""
    if user_id == 1:
        raise HTTPException(400, "User admin utama tidak dapat dihapus")

    try:
        with get_cursor() as cur:
            # Prevent deleting self? (Frontend should handle, or backend check if JWT existed)
            cur.execute("DELETE FROM users WHERE id = %s RETURNING id", (user_id,))
            deleted = cur.fetchone()
            if not deleted:
                raise HTTPException(404, "User tidak ditemukan")
            return {"success": True, "message": "User berhasil dihapus"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error deleting user: {e}")

@app.post("/admin/users/{user_id}/reset-password")
def reset_password(user_id: int):
    """Reset user password by Admin (auto-generated)"""
    # Proteksi: User ID 1 tidak boleh di-reset password-nya
    if user_id == 1:
        raise HTTPException(403, "Password Admin Utama (ID 1) tidak dapat di-reset")
    
    import random
    import string
    
    try:
        # Generate random 8-char password
        chars = string.ascii_letters + string.digits
        temp_password = ''.join(random.choice(chars) for _ in range(8))
        hashed_password = hash_password(temp_password)
        
        with get_cursor() as cur:
            # Check user existence
            cur.execute("SELECT id FROM users WHERE id = %s", (user_id,))
            if not cur.fetchone():
                raise HTTPException(404, "User tidak ditemukan")
            
            # Update password and set force flag
            cur.execute("""
                UPDATE users 
                SET password_hash = %s, force_password_change = true 
                WHERE id = %s
            """, (hashed_password, user_id))
            
            return {
                "success": True, 
                "message": "Password berhasil direset",
                "temporary_password": temp_password
            }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Reset password error: {e}")

@app.post("/auth/login")
def login_user(user: UserLogin):
    """Login user and return user info"""
    try:
        with get_cursor() as cur:
            password_hash = hash_password(user.password)
            cur.execute("""
                SELECT id, username, role, force_password_change FROM users 
                WHERE username = %s AND password_hash = %s
            """, (user.username, password_hash))
            
            found_user = cur.fetchone()
            if not found_user:
                raise HTTPException(401, "Username atau password salah")
            
            return {
                "success": True,
                "user": {
                    "id": found_user["id"],
                    "username": found_user["username"],
                    "role": found_user["role"],
                    "isAdmin": found_user["role"] == "admin",
                    "force_password_change": found_user.get("force_password_change", False)
                }
            }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Login error: {e}")

@app.put("/auth/profile")
def update_profile(user: UserUpdate):
    """Update user profile (username/password)"""
    try:
        with get_cursor() as cur:
            # 1. Verify current password
            password_hash = hash_password(user.current_password)
            cur.execute("""
                SELECT id, username, role FROM users 
                WHERE id = %s AND password_hash = %s
            """, (user.user_id, password_hash))
            
            current_user = cur.fetchone()
            if not current_user:
                raise HTTPException(401, "Password saat ini salah")
            
            updates = []
            params = []
            
            # 2. Prepare updates
            if user.username and user.username != current_user["username"]:
                # Check if new username exists
                cur.execute("SELECT id FROM users WHERE username = %s", (user.username,))
                if cur.fetchone():
                    raise HTTPException(400, "Username sudah digunakan")
                updates.append("username = %s")
                params.append(user.username)
            
            if user.new_password:
                new_hash = hash_password(user.new_password)
                updates.append("password_hash = %s")
                params.append(new_hash)
                updates.append("force_password_change = false")
            
            if not updates:
                return {
                    "success": True, 
                    "message": "Tidak ada perubahan",
                    "user": {
                        "id": current_user["id"],
                        "username": current_user["username"],
                        "role": current_user["role"],
                        "isAdmin": current_user["role"] == "admin"
                    }
                }
            
            # 3. Execute update
            query = f"UPDATE users SET {', '.join(updates)} WHERE id = %s RETURNING id, username, role"
            params.append(user.user_id)
            
            cur.execute(query, tuple(params))
            updated_user = cur.fetchone()
            
            return {
                "success": True,
                "user": {
                    "id": updated_user["id"],
                    "username": updated_user["username"],
                    "role": updated_user["role"],
                    "isAdmin": updated_user["role"] == "admin"
                }
            }
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Update error: {e}")

# Mapping sensor → nama tabel (hanya nama yang valid, mencegah SQL injection)
TABLES = {
    "dht22": "data_dht22",
    "mq2": "data_mq2",
    "pzem004t": "data_pzem004t",
    "bh1750": "data_bh1750"
}

# Mapping kolom per sensor (Updated to match DB schema)
COLUMNS = {
    "dht22": ["timestamp", "id", "temperature", "humidity"],
    "mq2": ["timestamp", "id", "gas_lpg", "gas_co", "smoke"],
    "pzem004t": ["timestamp", "id", "voltage", "current", "power", "energy", "power_factor"],
    "bh1750": ["timestamp", "id", "lux"]
}

# Range waktu dengan interval sampling optimal
RANGES = {
    "1h": {"delta": timedelta(hours=1), "interval": "10 minutes"},   # 6 points
    "6h": {"delta": timedelta(hours=6), "interval": "1 hour"},       # 6 points
    "12h": {"delta": timedelta(hours=12), "interval": "2 hours"},    # 6 points
    "24h": {"delta": timedelta(hours=24), "interval": "4 hours"},    # 6 points
    "7d": {"delta": timedelta(days=7), "interval": "1 day"},         # 7 points
}

def validate_sensor(sensor: str):
    """Validasi nama sensor dan return nama tabel yang aman"""
    if sensor not in TABLES:
        raise HTTPException(400, f"Sensor '{sensor}' tidak dikenal. Sensor yang tersedia: {list(TABLES.keys())}")
    return TABLES[sensor], COLUMNS[sensor]


# =========================================
# GET DATA TERBARU
# =========================================
@app.get("/latest/{sensor}")
def get_latest(sensor: str):
    """Mendapatkan data terbaru dari sensor"""
    table, columns = validate_sensor(sensor)

    try:
        with get_cursor() as cur:
            cur.execute(f"""
                SELECT * FROM {table}
                ORDER BY timestamp DESC
                LIMIT 1;
            """)
            row = cur.fetchone()

        if not row:
            return {"message": f"Belum ada data untuk sensor '{sensor}'"}

        return row

    except Exception as e:
        raise HTTPException(500, f"Database error: {e}")


# =========================================
# GET HISTORY (dengan sampling otomatis)
# =========================================
@app.get("/history/{sensor}")
def get_history(
    sensor: str, 
    range: str,
    raw: Optional[bool] = Query(False, description="Jika True, ambil semua data tanpa sampling")
):
    """
    Mendapatkan history data sensor dengan rentang waktu tertentu.
    
    - **sensor**: nama sensor (dht22, mq2, pzem004t, bh1750)
    - **range**: rentang waktu (1h, 6h, 12h, 24h, 7d)
    - **raw**: jika True, ambil semua data tanpa sampling (hati-hati data besar)
    """
    table, columns = validate_sensor(sensor)

    if range not in RANGES:
        raise HTTPException(
            400,
            f"Rentang waktu tidak valid. Pilihan: {list(RANGES.keys())}"
        )

    range_config = RANGES[range]
    time_limit = datetime.utcnow() - range_config["delta"]
    interval = range_config["interval"]

    try:
        with get_cursor() as cur:
            if interval and not raw:
                # Query dengan sampling menggunakan time_bucket (TimescaleDB) atau date_trunc
                # Menggunakan pendekatan yang kompatibel dengan PostgreSQL biasa
                numeric_cols = [c for c in columns if c not in ["id", "timestamp"]]
                avg_cols = ", ".join([f"AVG({col}) as {col}" for col in numeric_cols])
                
                cur.execute(f"""
                    SELECT 
                        to_timestamp(floor(extract(epoch from timestamp) / extract(epoch from interval '{interval}')) * extract(epoch from interval '{interval}'))
                        AS time_bucket,
                        {avg_cols},
                        COUNT(*) as sample_count
                    FROM {table}
                    WHERE timestamp >= %s
                    GROUP BY time_bucket
                    ORDER BY time_bucket ASC;
                """, (time_limit,))
            else:
                # Query tanpa sampling (untuk 1h atau jika raw=True)
                cur.execute(f"""
                    SELECT *
                    FROM {table}
                    WHERE timestamp >= %s
                    ORDER BY timestamp ASC;
                """, (time_limit,))

            rows = cur.fetchall()

        return {
            "sensor": sensor,
            "range": range,
            "sampled": bool(interval and not raw),
            "interval": interval if (interval and not raw) else None,
            "count": len(rows),
            "data": rows
        }

    except Exception as e:
        raise HTTPException(500, f"Database error: {e}")


# =========================================
# GET STATISTICS (agregasi cepat)
# =========================================
@app.get("/stats/{sensor}")
def get_stats(sensor: str, range: str):
    """
    Mendapatkan statistik agregasi dari sensor (min, max, avg).
    Sangat cepat karena hanya menghitung agregat.
    """
    table, columns = validate_sensor(sensor)

    if range not in RANGES:
        raise HTTPException(400, f"Rentang waktu tidak valid. Pilihan: {list(RANGES.keys())}")

    range_config = RANGES[range]
    time_limit = datetime.utcnow() - range_config["delta"]
    
    # Kolom numerik untuk agregasi
    numeric_cols = [c for c in columns if c not in ["id", "timestamp"]]
    
    # Build aggregation query
    agg_parts = []
    for col in numeric_cols:
        agg_parts.extend([
            f"MIN({col}) as {col}_min",
            f"MAX({col}) as {col}_max",
            f"AVG({col}) as {col}_avg"
        ])
    agg_query = ", ".join(agg_parts)

    try:
        with get_cursor() as cur:
            cur.execute(f"""
                SELECT 
                    COUNT(*) as total_records,
                    MIN(timestamp) as first_record,
                    MAX(timestamp) as last_record,
                    {agg_query}
                FROM {table}
                WHERE timestamp >= %s;
            """, (time_limit,))
            
            row = cur.fetchone()

        return {
            "sensor": sensor,
            "range": range,
            "stats": row
        }

    except Exception as e:
        raise HTTPException(500, f"Database error: {e}")


# =========================================
# EXPORT TO EXCEL
# =========================================
@app.get("/export/{sensor}")
def export_excel(sensor: str):
    """
    Download semua data history sensor dalam format Excel.
    """
    table, columns = validate_sensor(sensor)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{sensor.upper()} History"
        
        with get_cursor() as cur:
            # Fetch all data descending (newest first)
            col_list = ", ".join(columns)
            cur.execute(f"SELECT {col_list} FROM {table} ORDER BY timestamp DESC")
            
            # Use defined columns order
            col_names = columns
            
            # Style for header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
            
            # Write headers
            for col_idx, col_name in enumerate(col_names, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Write data rows
            row_num = 2
            while True:
                rows = cur.fetchmany(1000)
                if not rows:
                    break
                for row in rows:
                    for col_idx, col_name in enumerate(col_names, 1):
                        value = row[col_name]
                        # Format timestamp for better readability
                        if col_name == 'timestamp' and value:
                            value = value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'strftime') else str(value)
                        ws.cell(row=row_num, column=col_idx, value=value)
                    row_num += 1
            
            # Auto-adjust column widths
            for col_idx, col_name in enumerate(col_names, 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(col_name) + 5, 15)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{sensor}_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        raise HTTPException(500, f"Export error: {e}")


# =========================================
# SETTINGS ENDPOINTS
# =========================================
class SettingsUpdate(BaseModel):
    thresholds: dict
    enable_thresholds: Optional[bool] = True
    telegram_config: Optional[dict] = None

class TelegramTest(BaseModel):
    bot_token: str
    chat_id: str
    message: str = "Test notifikasi dari Smart Home Dashboard! 🚀"

@app.get("/settings")
def get_settings():
    """Get all application settings including thresholds"""
    try:
        with get_cursor() as cur:
            cur.execute("SELECT setting_key, setting_value FROM app_settings")
            rows = cur.fetchall()
            
            settings = {}
            for row in rows:
                settings[row['setting_key']] = row['setting_value']
            
            # Return default if no settings found
            if not settings.get('thresholds'):
                settings['thresholds'] = {
                    "dht22": {"tempMax": 35, "tempMin": 15, "humMax": 80, "humMin": 30},
                    "mq2": {"smokeMax": 500, "smokeWarn": 350, "lpgMax": 1000, "lpgWarn": 500, "coMax": 500, "coWarn": 200},
                    "pzem004t": {"powerMax": 2000, "voltageMin": 180, "voltageMax": 240, "currentMax": 10, "energyMax": 100, "pfMin": 0.85},
                    "bh1750": {"luxMax": 100000, "luxMin": 0}
                }
            
            # Get enable_thresholds setting (default to True if not found)
            if 'enable_thresholds' not in settings:
                settings['enable_thresholds'] = True
                
            # Get telegram_config setting (default dict if not found)
            if 'telegram_config' not in settings:
                settings['telegram_config'] = {"bot_token": "", "chat_id": "", "enabled": False}
            
            return {"success": True, "settings": settings}
    except Exception as e:
        raise HTTPException(500, f"Error fetching settings: {e}")


@app.put("/settings")
def update_settings(settings_data: SettingsUpdate):
    """Update application settings (thresholds)"""
    try:
        thresholds = settings_data.thresholds
        
        # Validate thresholds
        errors = []
        
        # DHT22 validation
        if 'dht22' in thresholds:
            dht = thresholds['dht22']
            if dht.get('tempMin') is not None and dht.get('tempMax') is not None:
                if dht['tempMin'] > dht['tempMax']:
                    errors.append("Suhu Min tidak boleh lebih besar dari Suhu Max")
            if dht.get('humMin') is not None and dht.get('humMax') is not None:
                if dht['humMin'] > dht['humMax']:
                    errors.append("Kelembaban Min tidak boleh lebih besar dari Kelembaban Max")
        
        # MQ2 validation
        if 'mq2' in thresholds:
            mq = thresholds['mq2']
            if mq.get('smokeWarn') is not None and mq.get('smokeMax') is not None:
                if mq['smokeWarn'] > mq['smokeMax']:
                    errors.append("Smoke Waspada tidak boleh lebih besar dari Smoke Bahaya")
        
        # PZEM004T validation
        if 'pzem004t' in thresholds:
            pz = thresholds['pzem004t']
            if pz.get('voltageMin') is not None and pz.get('voltageMax') is not None:
                if pz['voltageMin'] > pz['voltageMax']:
                    errors.append("Tegangan Min tidak boleh lebih besar dari Tegangan Max")
        
        # BH1750 validation
        if 'bh1750' in thresholds:
            bh = thresholds['bh1750']
            if bh.get('luxMin') is not None and bh.get('luxMax') is not None:
                if bh['luxMin'] > bh['luxMax']:
                    errors.append("Cahaya Min tidak boleh lebih besar dari Cahaya Max")
        
        if errors:
            raise HTTPException(400, "; ".join(errors))
            
        with get_cursor() as cur:
            import json
            
            # Update thresholds
            cur.execute(
                "INSERT INTO app_settings (setting_key, setting_value, updated_at) VALUES (%s, %s, NOW()) ON CONFLICT (setting_key) DO UPDATE SET setting_value = %s, updated_at = NOW()",
                ('thresholds', json.dumps(thresholds), json.dumps(thresholds))
            )
            
            # Update enable_thresholds if provided
            if settings_data.enable_thresholds is not None:
                cur.execute(
                    "INSERT INTO app_settings (setting_key, setting_value, updated_at) VALUES (%s, %s, NOW()) ON CONFLICT (setting_key) DO UPDATE SET setting_value = %s, updated_at = NOW()",
                    ('enable_thresholds', json.dumps(settings_data.enable_thresholds), json.dumps(settings_data.enable_thresholds))
                )
                
            # Update telegram_config if provided
            if settings_data.telegram_config is not None:
                cur.execute(
                    "INSERT INTO app_settings (setting_key, setting_value, updated_at) VALUES (%s, %s, NOW()) ON CONFLICT (setting_key) DO UPDATE SET setting_value = %s, updated_at = NOW()",
                    ('telegram_config', json.dumps(settings_data.telegram_config), json.dumps(settings_data.telegram_config))
                )
            
            # Fetch the updated thresholds to return
            cur.execute("SELECT setting_value FROM app_settings WHERE setting_key = 'thresholds'")
            result = cur.fetchone()
            
            return {
                "success": True, 
                "message": "Pengaturan berhasil disimpan",
                "thresholds": result['setting_value'] if result else thresholds,
                "enable_thresholds": settings_data.enable_thresholds,
                "telegram_config": settings_data.telegram_config
            }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error saving settings: {e}")

@app.post("/notify/telegram/test")
def test_telegram(data: TelegramTest):
    """Test send Telegram message"""
    import requests
    try:
        url = f"https://api.telegram.org/bot{data.bot_token}/sendMessage"
        
        # Simulasikan pesan alert jika diminta
        if "alert" in data.message.lower() or data.message == "test_alert":
            message = "⚠️ *SAMPLE ALERT (TEST)*\nSensor: *DHT22*\nKondisi: *Suhu Tinggi*\nNilai: *36.5 °C* (Batas: 35.0 °C)"
        else:
            message = f"🔔 *TEST KONEKSI SUCCESS*\n\n{data.message}\n\nJika Anda menerima pesan ini, berarti Bot Token dan Chat ID Anda sudah benar! ✅"

        payload = {
            "chat_id": data.chat_id,
            "text": message,
            "parse_mode": "Markdown"
        }
        res = requests.post(url, json=payload, timeout=5)
        if res.status_code == 200:
            return {"success": True, "message": "Pesan terkirim!"}
        else:
            return {"success": False, "message": f"Gagal: {res.text}"}
    except Exception as e:
        raise HTTPException(500, f"Error sending message: {e}")


@app.post("/settings/reset")
def reset_settings():
    """Reset settings to default values"""
    try:
        default_thresholds = {
            "dht22": {"tempMax": 35, "tempMin": 15, "humMax": 80, "humMin": 30},
            "mq2": {"smokeMax": 500, "smokeWarn": 350, "lpgMax": 1000, "lpgWarn": 500, "coMax": 500, "coWarn": 200},
            "pzem004t": {"powerMax": 2000, "voltageMin": 180, "voltageMax": 240, "currentMax": 10, "energyMax": 100, "pfMin": 0.85},
            "bh1750": {"luxMax": 100000, "luxMin": 0}
        }
        
        default_enable_thresholds = False
        default_telegram_config = {"bot_token": "", "chat_id": "", "enabled": False}

        with get_cursor() as cur:
            import json
            # Reset thresholds
            cur.execute("""
                INSERT INTO app_settings (setting_key, setting_value, updated_at)
                VALUES ('thresholds', %s, NOW())
                ON CONFLICT (setting_key) 
                DO UPDATE SET setting_value = EXCLUDED.setting_value, updated_at = NOW()
            """, (json.dumps(default_thresholds),))
            
            # Reset enable_thresholds to False
            cur.execute("""
                INSERT INTO app_settings (setting_key, setting_value, updated_at)
                VALUES ('enable_thresholds', %s, NOW())
                ON CONFLICT (setting_key) 
                DO UPDATE SET setting_value = EXCLUDED.setting_value, updated_at = NOW()
            """, (json.dumps(default_enable_thresholds),))
            
            # Reset telegram_config to disabled
            cur.execute("""
                INSERT INTO app_settings (setting_key, setting_value, updated_at)
                VALUES ('telegram_config', %s, NOW())
                ON CONFLICT (setting_key) 
                DO UPDATE SET setting_value = EXCLUDED.setting_value, updated_at = NOW()
            """, (json.dumps(default_telegram_config),))
            
            return {
                "success": True, 
                "message": "Pengaturan berhasil direset ke default (Notifikasi dinonaktifkan)",
                "thresholds": default_thresholds,
                "enable_thresholds": default_enable_thresholds,
                "telegram_config": default_telegram_config
            }
    except Exception as e:
        raise HTTPException(500, f"Error resetting settings: {e}")
